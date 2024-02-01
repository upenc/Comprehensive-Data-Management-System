from tkinter import *
import openpyxl
from tkinter import messagebox
from openpyxl import Workbook
from tkinter.ttk import Combobox
from openpyxl.styles import Alignment
import pathlib
from tkcalendar import DateEntry

root=Tk()
root.title("Bill Generator")
root.geometry('700x400+300+200')
root.resizable(False,True)
root.configure(bg="#326273")
file = pathlib.Path('Monthly_bill.xlsx')

if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A4'] = "Serial Number"
    sheet['B4']="Location"
    sheet['C4']="Date"
    sheet['D4']="Snacks @ Rs. 15/-"
    sheet['E4']="Lunch @ Rs. 50/-"
    sheet['F4']="Dinner @ Rs. 50/-"
    sheet['G4']="Tea @ Rs. 5/- 07/-"
    sheet['H4']="Request Of"
    sheet['I4']="Amount"

   # file.save('Monthly_bill.xlsx')
    sheet.merge_cells('A1:I1')
    sheet.merge_cells('A2:I2')
    sheet.merge_cells('A3:I3')
            
    sheet['A1'] = "FOODING BILLS - TRUCK PARKING YARD CANTEEN"
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    sheet['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    total_row_exists = any(row[0].value == "Total" for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row))
    # Saving the workbook
    file.save('Monthly_bill.xlsx')


def clear():
    Location_combobox.set('')
    DateValue_1.set('')
    SnacksValue.set('')
    LunchValue.set('')
    DinnerValue.set('')
    TeaValue.set('')
    RequestValue.set('')


serial_number = 1
def submit():
    global serial_number
    location = Location_combobox.get()
    date = DateValue_1.get()
    date_from = DateValue_From.get()
    date_to = DateValue_To.get()
    snacks = int(SnacksValue.get())
    lunch = int(LunchValue.get())
    dinner = int(DinnerValue.get())
    tea = int(TeaValue.get())
    request_of = RequestValue.get()
    file = openpyxl.load_workbook('Monthly_bill.xlsx')
    sheet = file.active
    #Serial number will increment from last serial
    last_serial = 0
    for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row, values_only=True):
        if row[0] is not None and str(row[0]).isdigit():
            last_serial = max(int(last_serial), int(row[0]))

    # Increment the serial number
    serial_number = last_serial + 1
    amount = snacks*15 + lunch*60 + dinner*60 + tea*7
    sheet.cell(column=1, row=sheet.max_row+1, value=serial_number)
    sheet.cell(column=2,row=sheet.max_row,value=location)
    sheet.cell(column=3,row=sheet.max_row,value=date)
    sheet.cell(column=4,row=sheet.max_row,value=snacks)
    sheet.cell(column=5,row=sheet.max_row,value=lunch)
    sheet.cell(column=6,row=sheet.max_row,value=dinner)
    sheet.cell(column=7,row=sheet.max_row,value=tea)
    sheet.cell(column=8,row=sheet.max_row,value=request_of)
    sheet.cell(column=9,row=sheet.max_row,value=amount)

    sheet['A2'] = f"(Period From ,{date_from} ,To,  {date_to})"
    file.save(r'Monthly_bill.xlsx')
    messagebox.showinfo("Submission Successful", "Data has been successfully submitted!")
    clear()

#icon
icon_image = PhotoImage(file="log.png")
root.iconphoto(False,icon_image)
#combobox for location
Location_combobox = Combobox(root,values=['Process','Machanical','VAT','Packing Plant','Railway'],width=43,font="arial 12")
Location_combobox.place(x=200,y=100)
#heading
Label(root,text="SK Canteen",font="Kis 13",bg="#326273",fg="#fff").place(x=300,y=20)
Label(root,text="From -",font="Kis 13",bg="#326273",fg="#fff").place(x=220,y=50)
Label(root,text="To -",font="Kis 13",bg="#326273",fg="#fff").place(x=400,y=50)
#label
Label(root,text="Location:",font=23,bg="#326273",fg="#fff").place(x=100,y=100)
Label(root,text="Date:",font=23,bg="#326273",fg="#fff").place(x=100,y=150)
Label(root,text="Snacks:",font=23,bg="#326273",fg="#fff").place(x=100,y=200)
Label(root,text="Lunch:",font=23,bg="#326273",fg="#fff").place(x=100,y=250)
Label(root,text="Dinner:",font=23,bg="#326273",fg="#fff").place(x=430,y=250)
Label(root,text="Tea:",font=23,bg="#326273",fg="#fff").place(x=430,y=200)
Label(root,text="Request of:",font=23,bg="#326273",fg="#fff").place(x=100,y=300)

#Entry
DateValue_1 = StringVar()
DateValue_From = StringVar()
DateValue_To = StringVar()
SnacksValue = StringVar()
LunchValue = StringVar()
DinnerValue = StringVar()
TeaValue = StringVar()
RequestValue = StringVar()

DateEntry(root, textvariable=DateValue_1, width=43, bd=2, font=20).place(x=200, y=150)
DateEntry(root, textvariable=DateValue_From, width=8, bd=2, font=20).place(x=270, y=50)
DateEntry(root, textvariable=DateValue_To, width=8, bd=2, font=20).place(x=430, y=50)
SnacksEntry = Entry(root,textvariable=SnacksValue,width=13,bd=2,font=20).place(x=200,y=200)
LunchEntry = Entry(root,textvariable=LunchValue,width=13,bd=2,font=20).place(x=200,y=250)
DinnerEntry = Entry(root,textvariable=DinnerValue,width=13,bd=2,font=20).place(x=487,y=250)
TeaEntry = Entry(root,textvariable=TeaValue,width=13,bd=2,font=20).place(x=487,y=200)
RequestEntry = Entry(root,textvariable=RequestValue,width=45,bd=2,font=20).place(x=200,y=300)


Button(root,text="Submit",bg="#326273",fg="white",width=12,height=2,command=submit).place(x=260,y=350)
Button(root,text="Clear",bg="#326273",fg="white",width=12,height=2,command=clear).place(x=380,y=350)
Button(root,text="Exit",bg="#326273",fg="white",width=12,height=2,command=lambda:root.destroy()).place(x=500,y=350)


root.mainloop()